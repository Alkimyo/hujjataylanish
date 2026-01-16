from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, FileResponse, StreamingHttpResponse
from django.views.decorators.http import require_http_methods
from django.core.exceptions import ValidationError, PermissionDenied
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.db.models import Q
from django.db import transaction
from django.utils import timezone

from .models import Document, DocumentType, User, Faculty, Department, Program, Notification, TeachingAllocation, Subject, Group, AcademicYear, University, AuditLog, JobRun
from .services import ApprovalWorkflowService, NotificationService, DocumentFilterService
from .qr_service import QRCodeService
from .forms import DocumentUploadForm, ProfileUpdateForm, PasswordChangeUzForm, SubjectImportForm, AllocationImportForm
import os
from .qr_service import QRCodeService
import json 
from django.core.serializers.json import DjangoJSONEncoder
import time
import csv
from openpyxl import load_workbook


# documents/views.py
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import json

@require_POST
@login_required
def switch_role(request):
    """Foydalanuvchi ro'lini o'zgartirish"""
    try:
        data = json.loads(request.body)
        role_code = data.get('role_code')
        role_name = data.get('role_name')
        
        # Foydalanuvchining rollaridan berilgan rol nomini topish
        # Faol rolni yangilash
        from .models import Role
        role_obj = None
        if role_code:
            role_obj = Role.get_role_by_code(role_code)
        if not role_obj and role_name:
            role_obj = Role.objects.filter(name__iexact=role_name, is_active=True).first()
        if not role_obj:
            role_type_lookup = {
                key: value for key, value in Role.ROLE_TYPE_CHOICES
            }
            role_type = None
            for key, value in role_type_lookup.items():
                if role_name.lower() == value.lower():
                    role_type = key
                    break
            if not role_type and role_name in role_type_lookup:
                role_type = role_name
            if role_type:
                role_obj = Role.get_default_role_for_type(role_type)

        if not role_obj:
            return JsonResponse({
                'success': False,
                'error': 'Rol topilmadi'
            })

        if not (request.user.has_role_type(role_obj.role_type) or request.user.has_role(role_obj.code)):
            return JsonResponse({
                'success': False,
                'error': 'Bu rol sizda mavjud emas'
            })

        # Rolni session'ga saqlash
        request.session['active_role_code'] = role_obj.code
        request.session['current_role'] = role_obj.name
        request.user.active_role = role_obj
        request.user.save(update_fields=['active_role'])

        AuditLog.objects.create(
            user=request.user,
            action='role_switched',
            metadata={
                'role_code': role_obj.code,
                'role_name': role_obj.name,
            },
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )

        return JsonResponse({
            'success': True,
            'role_name': role_obj.name,
            'message': 'Rol muvaffaqiyatli o\'zgartirildi'
        })
        
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def dashboard(request):
    user = request.user
    pending_approvals = ApprovalWorkflowService.get_pending_approvals_for_user(user)
    my_documents = Document.objects.filter(uploaded_by=user).order_by('-uploaded_at')[:10]
    notifications = Notification.objects.filter(recipient=user).order_by('-created_at')[:10]
    
    stats = {
        'pending_approvals': pending_approvals.count(),
        'my_documents_total': Document.objects.filter(uploaded_by=user).count(),
        'my_approved': Document.objects.filter(uploaded_by=user, status='approved').count(),
        'my_pending': Document.objects.filter(uploaded_by=user, status='pending_approval').count(),
        'my_rejected': Document.objects.filter(uploaded_by=user, status='rejected').count(),
        'unread_notifications': NotificationService.get_unread_count(user),
    }
    
    context = {
        'pending_approvals': pending_approvals,
        'my_documents': my_documents,
        'notifications': notifications,
        'stats': stats,
    }
    if request.user.has_role_type('department_head') and request.user.managed_department:
        return department_head_dashboard(request)
    else:
        return render(request, 'documents/dashboard.html', context)

@login_required
def pending_approvals(request):
    """Foydalanuvchi uchun kutilayotgan barcha tasdiqlar"""
    pending = ApprovalWorkflowService.get_pending_approvals_for_user(request.user)
    context = {'pending_approvals': pending}
    return render(request, 'documents/pending_approvals.html', context)

# --- APPROVAL ACTIONS (ENG MUHIM QISM) ---

@login_required
@require_http_methods(["POST"])
def approve_document(request, document_id):
    """Hujjatni tasdiqlash"""
    try:
        # HTMLda tasdiqlash uchun comment input yo'q, shuning uchun bo'sh string olamiz
        # Agar kelajakda qo'shilsa, request.POST.get('comment', '') ishlayveradi
        comment = request.POST.get('comment', '')
        
        result = ApprovalWorkflowService.approve_document(
            document_id=document_id,
            user=request.user,
            comment=comment,
            request=request
        )
        
        messages.success(request, result['message'])
        
    except (ValidationError, PermissionDenied) as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Xatolik yuz berdi: {str(e)}")
        
    # Qayerdan kelgan bo'lsa o'sha yerga qaytish (dashboard yoki pending list)
    return redirect(request.META.get('HTTP_REFERER', 'pending_approvals'))


@login_required
@require_http_methods(["POST"])
def reject_document(request, document_id):
    """Hujjatni rad etish"""
    try:
        # DIQQAT: HTML formadagi textarea name="comment" edi.
        # Services da esa reason deb nomlaganmiz.
        # Bu yerda mapping qilamiz.
        reason = request.POST.get('comment', '') # HTML name='comment'
        
        if not reason.strip():
            messages.error(request, "Rad etish sababini yozish shart!")
            return redirect(request.META.get('HTTP_REFERER', 'pending_approvals'))
        
        result = ApprovalWorkflowService.reject_document(
            document_id=document_id,
            user=request.user,
            reason=reason,
            request=request
        )
        
        messages.warning(request, result['message'])
        
    except (ValidationError, PermissionDenied) as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Xatolik yuz berdi: {str(e)}")
        
    return redirect(request.META.get('HTTP_REFERER', 'pending_approvals'))


@login_required
def document_list(request):
    user = request.user
    
    if user.role in ['admin', 'director', 'director_deputy']:
        documents = Document.objects.all()
        
    elif user.role in ['faculty_dean', 'dean_deputy']:
        if hasattr(user, 'managed_faculty') and user.managed_faculty:
             documents = Document.objects.filter(uploaded_by__faculty=user.managed_faculty)
        elif user.faculty:
             documents = Document.objects.filter(uploaded_by__faculty=user.faculty)
        else:
             documents = Document.objects.filter(uploaded_by=user)

    elif user.role == 'department_head':
        if hasattr(user, 'managed_department') and user.managed_department:
             documents = Document.objects.filter(uploaded_by__department=user.managed_department)
        elif user.department:
             documents = Document.objects.filter(uploaded_by__department=user.department)
        else:
             documents = Document.objects.filter(uploaded_by=user)

    else:
        documents = Document.objects.filter(uploaded_by=user)

    context = {
        'document_types': DocumentType.objects.filter(is_active=True),
        'subjects': Subject.objects.all(),
        'academic_years': AcademicYear.objects.filter(is_active=True),
        'programs': Program.objects.all(),
        'filters': request.GET
    }

    if user.role in ['department_head', 'faculty_dean', 'dean_deputy', 'director', 'director_deputy', 'admin']:
        context['can_filter_author'] = True

    if user.role in ['faculty_dean', 'dean_deputy', 'director', 'director_deputy', 'admin']:
        if user.role in ['faculty_dean', 'dean_deputy'] and hasattr(user, 'managed_faculty'):
             context['departments'] = Department.objects.filter(faculty=user.managed_faculty)
        else:
             context['departments'] = Department.objects.all()

    if user.role in ['director', 'director_deputy', 'admin']:
        context['faculties'] = Faculty.objects.all()

    if user.role == 'admin':
        context['universities'] = University.objects.all()

    status = request.GET.get('status')
    doc_type = request.GET.get('document_type')
    program = request.GET.get('program')
    subject = request.GET.get('subject')
    year = request.GET.get('academic_year')
    author = request.GET.get('author')
    department = request.GET.get('department')
    faculty = request.GET.get('faculty')
    university = request.GET.get('university')

    if status: documents = documents.filter(status=status)
    if doc_type: documents = documents.filter(document_type_id=doc_type)
    if program: documents = documents.filter(related_group__program_id=program)
    if subject: documents = documents.filter(subject_id=subject)
    if year: documents = documents.filter(academic_year_id=year)
    
    if author and context.get('can_filter_author'):
        documents = documents.filter(
            Q(uploaded_by__first_name__icontains=author) | 
            Q(uploaded_by__last_name__icontains=author) |
            Q(uploaded_by__username__icontains=author)
        )
    
    # Tashkiliy tuzilma bo'yicha filterlar (faqat ruxsati borlarga)
    if department and context.get('departments'):
        documents = documents.filter(uploaded_by__department_id=department)
        
    if faculty and context.get('faculties'):
        documents = documents.filter(uploaded_by__faculty_id=faculty)

    if university and context.get('universities'):
        documents = documents.filter(uploaded_by__university_id=university)
        

    context['documents'] = documents.order_by('-uploaded_at')

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'documents/_document_table.html', context)

    return render(request, 'documents/document_list.html', context)
    

@login_required
def document_detail(request, document_id):
    # Optimize qilingan query (barcha bog'liqliklar bilan)
    document = get_object_or_404(
        Document.objects.select_related(
            'document_type', 
            'uploaded_by', 
            'uploaded_by__department', 
            'uploaded_by__faculty',
            'subject',
            'related_group'
        ), 
        id=document_id
    )
    
    if not request.user.can_view_document(document):
        raise PermissionDenied("Hujjatni ko'rishga ruxsat yo'q")
    
    # Workflow tarixi
    workflow_info = ApprovalWorkflowService.get_document_history(document)
    
    # Tasdiqlash huquqi bormi?
    current_step = document.get_current_approver()
    can_approve = False
    if current_step and current_step.status == 'pending':
        if current_step.approver == request.user:
            can_approve = True

    # --- MA'LUMOTLARNI ALOHIDA O'ZGARUVCHILARGA OLISH ---
    # Bu usul xavfsizroq va shablonda logic kamayadi
    
    doc_title = document.title
    doc_status_display = document.get_status_display()
    doc_status_code = document.status
    
    # Hujjat turi nomi (xatolik shu yerda bo'lishi mumkin edi)
    doc_type_name = document.document_type.name if document.document_type else "Noma'lum tur"
    
    # Yuklagan shaxs
    uploader_full_name = document.uploaded_by.middle_name + " " + document.uploaded_by.first_name + " " + document.uploaded_by.last_name if document.uploaded_by else "Noma'lum"
    uploader_role = document.uploaded_by.get_role_display()
    
    # Sana formatlash
    uploaded_at = document.uploaded_at
    
    # Fayl ma'lumotlari
    file_name = document.file_name
    file_size = document.file_size
    verification_code = document.verification_code
    description = document.description
    
    # Qo'shimcha ma'lumotlar (agar bo'lmasa tire qo'yamiz)
    faculty_name = document.uploaded_by.faculty.name if document.uploaded_by.faculty else "-"
    department_name = document.uploaded_by.department.name if document.uploaded_by.department else "-"
    
    # Fan va Guruh (agar bo'lsa)
    subject_name = document.subject.name if document.subject else None
    group_name = document.related_group.name if document.related_group else None
    
    # Bosqich matni
    current_step_text = document.get_expected_approver_text()
    
    context = {
        'document': document, # Asosiy obyekt (ID va URL lar uchun kerak)
        
        # Alohida o'zgaruvchilar (Display uchun)
        'doc_title': doc_title,
        'doc_status_display': doc_status_display,
        'doc_status_code': doc_status_code,
        'doc_type_name': doc_type_name,
        'uploader_full_name': uploader_full_name,
        'uploader_role': uploader_role,
        'uploaded_at': uploaded_at,
        'file_name': file_name,
        'file_size': file_size,
        'verification_code': verification_code,
        'description': description,
        'faculty_name': faculty_name,
        'department_name': department_name,
        'subject_name': subject_name,
        'group_name': group_name,
        'current_step_text': current_step_text,
        
        'workflow_info': workflow_info,
        'can_approve': can_approve,
    }
    return render(request, 'documents/document_detail.html', context)

    
from django.core.serializers.json import DjangoJSONEncoder
import json

@login_required
@require_http_methods(["GET", "POST"])
def upload_document(request):
    if request.method == 'POST':
        form = DocumentUploadForm(request.POST, request.FILES, user=request.user)
        
        if form.is_valid():
            document = form.save(commit=False)
            document.uploaded_by = request.user
            document.file_name = request.FILES['file'].name
            document.file_size = request.FILES['file'].size
            
            # Qo'shimcha maydonlarni saqlash
            document.subject = form.cleaned_data.get('subject')
            document.academic_year = form.cleaned_data.get('academic_year')
            document.related_group = form.cleaned_data.get('related_group')
            
            document.save()
            
            messages.success(
                request, 
                f"✓ Hujjat muvaffaqiyatli yuklandi va tasdiqlashga yuborildi. "
                f"Tasdiqlash kodi: {document.verification_code}"
            )
            return redirect('document_detail', document_id=document.id)
        else:
            messages.error(request, "Iltimos, formadagi xatolarni tuzating.")
    else:
        form = DocumentUploadForm(user=request.user)
    
    # Foydalanuvchi yuklashi mumkin bo'lgan hujjat turlari
    available_doc_types = []
    for doc_type in DocumentType.objects.filter(is_active=True):
        if doc_type.can_user_upload(request.user):
            # Workflow to'g'ri formatlash
            workflow_display = doc_type.get_workflow_display()
            if workflow_display is None:
                workflow_display = ""

            available_doc_types.append({
                'id': doc_type.id,
                'name': doc_type.name,
                'max_size': float(doc_type.max_file_size_mb),
                'requires_subject': doc_type.requires_subject,
                'requires_academic_year': doc_type.requires_academic_year,
                'requires_group': doc_type.requires_group,
                'allowed_extensions': doc_type.allowed_extensions,
                'workflow': doc_type.get_workflow_display() or "",
            })
    # To'g'ri JSON formatlash
    import json
    from django.core.serializers.json import DjangoJSONEncoder
    
    available_doc_types_json = json.dumps(
        available_doc_types, 
        cls=DjangoJSONEncoder, 
        ensure_ascii=False
    )
    
    context = {
        'form': form,
        'available_doc_types_json': available_doc_types_json,
        'available_doc_types_list': available_doc_types,  # Shuning uchun ham qo'shdik
    }

    return render(request, 'documents/upload.html', context)
            


@login_required
def get_document_type_info(request, doc_type_id):
    """AJAX: Hujjat turi haqida ma'lumot qaytarish"""
    try:
        doc_type = DocumentType.objects.get(id=doc_type_id, is_active=True)
        
        # Foydalanuvchi bu turda hujjat yuklashi mumkinmi?
        if not doc_type.can_user_upload(request.user):
            return JsonResponse({
                'error': 'Sizning rolingiz bu hujjat turini yuklashga ruxsati yo\'q'
            }, status=403)
        
        return JsonResponse({
            'success': True,
            'max_size_mb': float(doc_type.max_file_size_mb),
            'allowed_extensions': doc_type.allowed_extensions,
            'requires_subject': doc_type.requires_subject,
            'requires_academic_year': doc_type.requires_academic_year,
            'requires_group': doc_type.requires_group,
            'workflow': doc_type.get_workflow_display(),
        })
    except DocumentType.DoesNotExist:
        return JsonResponse({'error': 'Hujjat turi topilmadi'}, status=404)

@login_required
def download_document(request, document_id):
    document = get_object_or_404(Document, id=document_id)
    if not request.user.can_view_document(document):
        raise PermissionDenied("Ruxsat yo'q")
    
    if document.status == 'approved' and document.final_pdf:
        file_path = document.final_pdf.path
        filename = f"approved_{document.file_name}"
    else:
        file_path = document.file.path
        filename = document.file_name
    
    try:
        response = FileResponse(open(file_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except FileNotFoundError:
        raise Http404("Fayl topilmadi")

import os
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import FileResponse, Http404
from django.contrib import messages

@login_required
def download_qr_code(request, document_id):
    """
    Bu funksiya endi faqat QR rasmni emas, balki 
    QR kod va rasmiy matn qo'shilgan to'liq PDF faylni yuklab beradi.
    """
    document = get_object_or_404(Document, id=document_id)
    
    # 1. Ruxsatni tekshirish
    if not request.user.can_view_document(document):
        raise PermissionDenied("Hujjatni ko'rishga ruxsat yo'q")
    
    # 2. Hujjat tasdiqlangan bo'lishi shart
    if document.status != 'approved':
        messages.warning(request, "Hujjat hali to'liq tasdiqlanmagan. QR kodli PDF faqat tasdiqlangandan so'ng shakllanadi.")
        return redirect('document_detail', document_id=document.id)
    
    # 3. Agar final_pdf (QR kodli versiya) hali yaratilmagan bo'lsa yoki fayl o'chib ketgan bo'lsa
    if not document.final_pdf or not os.path.exists(document.final_pdf.path):
        try:
            # Yangi PDF generatsiya qilamiz (Bu funksiya qr_service.py da yozilgan)
            QRCodeService.generate_final_pdf(document)
        except Exception as e:
            messages.error(request, f"PDF generatsiya qilishda xatolik yuz berdi: {str(e)}")
            return redirect('document_detail', document_id=document.id)
    
    # 4. Faylni foydalanuvchiga jo'natish
    try:
        file_path = document.final_pdf.path
        
        # Faylni ochib response ga beramiz
        response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
        
        # Fayl nomini chiroyli qilish (masalan: approved_HujjatNomi.pdf)
        filename = f"approved_{document.file_name}"
        if not filename.lower().endswith('.pdf'):
            filename += '.pdf'
            
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except FileNotFoundError:
        raise Http404("Tasdiqlangan PDF fayl serverda topilmadi.")
# --- OTHER VIEWS ---

@login_required
def profile_view(request):
    user = request.user
    profile_form = ProfileUpdateForm(instance=user)
    password_form = PasswordChangeUzForm(user)

    if request.method == 'POST':
        if 'update_profile' in request.POST:
            profile_form = ProfileUpdateForm(request.POST, instance=user)
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, "Profil yangilandi")
                return redirect('profile')

        elif 'change_password' in request.POST:
            password_form = PasswordChangeUzForm(user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Parol o‘zgartirildi")
                return redirect('profile')

    context = {
        'profile_form': profile_form,
        'password_form': password_form,
    }
    return render(request, 'documents/profile.html', context)

@login_required
def subject_distribution(request):
    """Fan taqsimotlari"""
    
    # Foydalanuvchi tekshiruvi
    if request.user.role != 'department_head':
        messages.error(request, "Huquqingiz yo'q")
        return redirect('dashboard')
    
    # Kafedra tekshiruvi
    if not hasattr(request.user, 'managed_department') or not request.user.managed_department:
        messages.error(request, "Kafedra birigtirilmagan")
        return redirect('dashboard')
    
    department = request.user.managed_department
    
    # Form yaratish - DIQQAT: department=department parametrini berish
    if request.method == 'POST':
        # Formaga department parametrini berish
        form = TeachingAllocationForm(request.POST, department=department)
        
        if form.is_valid():
            allocation = form.save(commit=False)
            allocation.department = department
            allocation.created_by = request.user
            allocation.save()
            
            messages.success(
                request, 
                f"{allocation.teacher.get_full_name()} ga "
                f"'{allocation.subject.name}' fani biriktirildi."
            )
            return redirect('subject_distribution')
    else:
        # GET so'rovi uchun ham department parametrini berish
        form = TeachingAllocationForm(department=department)
    
    # Taqsimotlarni olish
    allocations = TeachingAllocation.objects.filter(
        department=department
    ).select_related(
        'teacher', 'subject', 'group', 'academic_year'
    ).order_by('-created_at')
    
    context = {
        'form': form,
        'allocations': allocations,
        'department': department
    }
    
    return render(request, 'documents/subject_distribution.html', context)

@login_required
def notifications_list(request):
    """View all notifications with document workflow details"""
    
    # 1. Bildirishnomalarni olamiz
    notifications = Notification.objects.filter(
        recipient=request.user
    ).select_related('document', 'document__document_type').order_by('-created_at')
    
    # 2. Har bir bildirishnoma uchun hujjat jarayonini tayyorlaymiz
    notifications_with_workflow = []
    
    for notif in notifications:
        workflow_steps = []
        if notif.document:
            # Hujjatning barcha bosqichlarini olamiz
            steps = notif.document.approval_steps.all().order_by('step_order')
            
            for step in steps:
                workflow_steps.append({
                    'role': step.role_required, # Yoki get_role_required_display()
                    'status': step.status, # pending, approved, rejected
                    'approver_name': step.approver.get_full_name() if step.approver else "Noma'lum",
                    'date': step.approved_at
                })
        
        notif.workflow_data = workflow_steps
        notifications_with_workflow.append(notif)

    context = {
        'notifications': notifications_with_workflow,
    }
    
    return render(request, 'documents/notifications.html', context)

@login_required
@require_http_methods(["POST"])
def mark_notification_read(request, notification_id):
    NotificationService.mark_as_read(notification_id, request.user)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    return redirect('notifications_list')

@login_required
@require_http_methods(["POST"])
def mark_all_notifications_read(request):
    NotificationService.mark_all_as_read(request.user)
    return redirect('notifications_list')

# Public Verify Views
def verify_document(request):
    # Logic for public verification form
    return render(request, 'documents/verify.html')

from django.http import FileResponse
from django.shortcuts import render
from .models import Document

def verify_by_uuid(request, uuid):
    """
    Kod kiritilganda:
    1. Agar to'g'ri bo'lsa -> Avtomatik fayl yuklanadi.
    2. Agar xato bo'lsa -> Xato xabari chiqadi.
    """
    error = None

    def _serve_verified_document(document):
        target_file = document.final_pdf if document.final_pdf else document.file
        if os.path.exists(target_file.path):
            response = FileResponse(open(target_file.path, 'rb'), content_type='application/pdf')
            filename = f"hujjat_{document.verification_code}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        raise FileNotFoundError("Hujjat bazada bor, lekin fayl serverda topilmadi.")

    document = Document.objects.filter(uuid=uuid, status='approved').first()
    if document and request.method == 'GET':
        try:
            return _serve_verified_document(document)
        except Exception as e:
            error = f"Faylni yuklashda xatolik: {str(e)}"

    if request.method == 'POST':
        code = request.POST.get('verification_code', '').strip().upper()
        document = Document.objects.filter(verification_code=code, status='approved').first()

        if document:
            try:
                return _serve_verified_document(document)
            except Exception as e:
                error = f"Faylni yuklashda xatolik: {str(e)}"
        else:
            error = "Bunday kodli hujjat topilmadi yoki hali tasdiqlanmagan."
    elif not document:
        error = "Bunday UUID li hujjat topilmadi yoki hali tasdiqlanmagan."

    return render(request, 'documents/verify.html', {'error': error})

# API Views
@login_required
def api_document_status(request, document_id):
    document = get_object_or_404(Document, id=document_id)
    # ... logic ...
    return JsonResponse({'status': document.status})

@login_required
def api_notification_count(request):
    count = NotificationService.get_unread_count(request.user)
    return JsonResponse({'unread_count': count})


@login_required
def api_notification_stream(request):
    def event_stream():
        last_count = None
        poll_seconds = 30

        while True:
            count = NotificationService.get_unread_count(request.user)
            if count != last_count:
                payload = json.dumps({'unread_count': count})
                yield f"event: count\ndata: {payload}\n\n"
                last_count = count

            yield ": keep-alive\n\n"
            time.sleep(poll_seconds)

    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response


@login_required
def jobs_health(request):
    if not request.user.is_staff:
        return JsonResponse({'error': 'Forbidden'}, status=403)

    jobs = JobRun.objects.all().order_by('task_name')
    data = []
    for job in jobs:
        data.append({
            'task_name': job.task_name,
            'last_status': job.last_status,
            'last_run_at': job.last_run_at.isoformat() if job.last_run_at else None,
            'last_success_at': job.last_success_at.isoformat() if job.last_success_at else None,
            'last_duration_ms': job.last_duration_ms,
            'last_error': job.last_error,
        })

    return JsonResponse({'jobs': data})









from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
from .models import Subject, TeachingAllocation, Department, Group
from .forms import SubjectForm, TeachingAllocationForm


def _normalize_header(value):
    text = str(value or '').strip().lower()
    text = text.replace(' ', '_').replace('-', '_')
    text = text.replace("'", "").replace("’", "").replace("`", "")
    return text


def _parse_import_file(uploaded_file):
    name = uploaded_file.name.lower()
    rows = []

    if name.endswith('.csv'):
        content = uploaded_file.read().decode('utf-8-sig', errors='ignore').splitlines()
        reader = csv.DictReader(content)
        for raw in reader:
            rows.append({k: v for k, v in raw.items()})
        return rows

    if name.endswith('.xlsx'):
        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active
        headers = []
        for cell in sheet[1]:
            headers.append(_normalize_header(cell.value))
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            row_data = {}
            for idx, value in enumerate(row):
                key = headers[idx] if idx < len(headers) else f'col_{idx}'
                row_data[key] = value
            rows.append(row_data)
        return rows

    raise ValueError("Faqat .csv yoki .xlsx fayllari qabul qilinadi.")


def _map_row_keys(row, alias_map):
    mapped = {}
    for key, value in row.items():
        norm_key = _normalize_header(key)
        canonical = alias_map.get(norm_key)
        if canonical:
            mapped[canonical] = value
    return mapped

@login_required
def department_head_dashboard(request):
    """Kafedra mudiri asosiy sahifasi"""
    
    # Faqat kafedra mudiri kirishi mumkin
    if (not request.user.has_role_type('department_head')) or (not request.user.managed_department):
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    department = request.user.managed_department
    
    # Statistika
    stats = {
        'subjects_count': Subject.objects.filter(department=department).count(),
        'teachers_count': User.objects.filter(active_role__role_type='teacher', department=department).count(),
        'allocations_count': TeachingAllocation.objects.filter(department=department).count(),
        'active_year': AcademicYear.objects.filter(is_active=True).first(),
    }
    
    # So'nggi qo'shilgan fanlar
    recent_subjects = Subject.objects.filter(
        department=department
    ).order_by('-created_at')[:5]
    
    # So'nggi taqsimotlar
    recent_allocations = TeachingAllocation.objects.filter(
        department=department
    ).select_related('teacher', 'subject', 'group', 'academic_year').order_by('-created_at')[:5]
    
    context = {
        'department': department,
        'stats': stats,
        'recent_subjects': recent_subjects,
        'recent_allocations': recent_allocations,
    }
    
    return render(request, 'department_head/dashboard.html', context)



@login_required
def subjects_list(request):
    """Fanlar ro'yxati"""
    
    if request.user.role != 'department_head' or not request.user.managed_department:
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    department = request.user.managed_department
    
    # Asosiy queryset - allocations_count ni annotate qilamiz
    from django.db.models import Count
    
    subjects = Subject.objects.filter(
        department=department
    ).annotate(
        allocations_count=Count('allocations')  # TeachingAllocation bilan bog'lanish
    ).order_by('name')
    
    # Qidiruv
    search_query = request.GET.get('search', '')
    if search_query:
        subjects = subjects.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Yo'nalish bo'yicha filtr
    program_filter = request.GET.get('program')
    if program_filter:
        subjects = subjects.filter(
            taught_in_programs__icontains=program_filter
        )
    
    # Umumiy son
    total_count = subjects.count()
    
    # Pagination
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    page = request.GET.get('page', 1)
    paginator = Paginator(subjects, 15)  # Har sahifada 15 ta
    
    try:
        subjects_page = paginator.page(page)
    except PageNotAnInteger:
        subjects_page = paginator.page(1)
    except EmptyPage:
        subjects_page = paginator.page(paginator.num_pages)
    
    # Kafedradagi yo'nalishlar (filtr uchun)
    programs = Program.objects.filter(
        department=department
    ).order_by('code')

    import_form = SubjectImportForm()
    
    context = {
        'subjects': subjects_page,  # Pagination object
        'programs': programs,
        'department': department,
        'search_query': search_query,
        'total_count': total_count,
        'import_form': import_form,
    }
    
    return render(request, 'department_head/subjects_list.html', context)


@login_required
def subjects_import(request):
    """Fanlarni CSV/XLSX orqali yuklash"""
    if request.user.role != 'department_head' or not request.user.managed_department:
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')

    department = request.user.managed_department
    if request.method != 'POST':
        return redirect('subjects_list')

    form = SubjectImportForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Fayl xato: " + "; ".join(form.errors.get('file', [])))
        return redirect('subjects_list')

    alias_map = {
        'name': 'name',
        'fan_nomi': 'name',
        'fan': 'name',
        'nomi': 'name',
        'subject_name': 'name',
        'code': 'code',
        'fan_kodi': 'code',
        'kodi': 'code',
        'subject_code': 'code',
        'credits': 'credits',
        'kredit': 'credits',
        'lecture_hours': 'lecture_hours',
        'maruza_soatlari': 'lecture_hours',
        'maruza': 'lecture_hours',
        'practice_hours': 'practice_hours',
        'amaliyot_soatlari': 'practice_hours',
        'amaliyot': 'practice_hours',
        'taught_in_programs': 'taught_in_programs',
        'yonalishlar': 'taught_in_programs',
        'yonalish_kodlari': 'taught_in_programs',
        'programs': 'taught_in_programs',
        'program_codes': 'taught_in_programs',
    }

    try:
        rows = _parse_import_file(form.cleaned_data['file'])
    except Exception as e:
        messages.error(request, f"Faylni o'qishda xatolik: {str(e)}")
        return redirect('subjects_list')

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        data = _map_row_keys(row, alias_map)
        name = str(data.get('name') or '').strip()
        code = str(data.get('code') or '').strip()

        if not name or not code:
            skipped += 1
            errors.append(f"{idx}-qator: fan nomi yoki kodi yo'q.")
            continue

        subject = Subject.objects.filter(code=code, department=department).first()
        is_new = subject is None
        if is_new:
            subject = Subject(code=code, department=department)

        subject.name = name

        credits = data.get('credits')
        if credits not in (None, ''):
            try:
                subject.credits = int(credits)
            except Exception:
                errors.append(f"{idx}-qator: kredit noto'g'ri qiymat.")

        lecture_hours = data.get('lecture_hours')
        if lecture_hours not in (None, ''):
            try:
                subject.lecture_hours = int(lecture_hours)
            except Exception:
                errors.append(f"{idx}-qator: ma'ruza soati noto'g'ri qiymat.")

        practice_hours = data.get('practice_hours')
        if practice_hours not in (None, ''):
            try:
                subject.practice_hours = int(practice_hours)
            except Exception:
                errors.append(f"{idx}-qator: amaliyot soati noto'g'ri qiymat.")

        programs = data.get('taught_in_programs')
        if programs not in (None, ''):
            subject.taught_in_programs = str(programs).strip()

        try:
            subject.save()
            if is_new:
                created += 1
            else:
                updated += 1
        except Exception as e:
            skipped += 1
            errors.append(f"{idx}-qator: saqlashda xatolik: {str(e)}")

    message = f"Import yakunlandi. Yangi: {created}, yangilangan: {updated}, o'tkazib yuborilgan: {skipped}."
    if errors:
        message += " Xatolar: " + " | ".join(errors[:10])
    messages.info(request, message)
    return redirect('subjects_list')
    
    

@login_required
def subject_add(request):
    """Yangi fan qo'shish"""
    
    if request.user.role != 'department_head' or not request.user.managed_department:
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    department = request.user.managed_department
    
    if request.method == 'POST':
        form = SubjectForm(request.POST, department=department)
        
        if form.is_valid():
            subject = form.save(commit=False)
            subject.department = department
            subject.save()
            
            messages.success(
                request, 
                f"✓ '{subject.name}' fani qo'shildi.\n"
                f"Yo'nalishlar: {subject.programs_display}"
            )
            return redirect('subjects_list')
    else:
        form = SubjectForm(department=department)
    
    context = {
        'form': form,
        'department': department,
        'title': 'Yangi fan qo\'shish',
        'is_edit': False,
    }
    
    return render(request, 'department_head/subject_form.html', context)


@login_required
def subject_edit(request, subject_id):
    """Fanni tahrirlash"""
    
    if request.user.role != 'department_head' or not request.user.managed_department:
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    department = request.user.managed_department
    subject = get_object_or_404(Subject, id=subject_id, department=department)
    
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject, department=department)
        
        if form.is_valid():
            form.save()
            messages.success(request, f"✓ '{subject.name}' fani yangilandi.")
            return redirect('subjects_list')
    else:
        form = SubjectForm(instance=subject, department=department)
    
    context = {
        'form': form,
        'subject': subject,
        'department': department,
        'title': 'Fanni tahrirlash',
        'is_edit': True,
    }
    
    return render(request, 'department_head/subject_form.html', context)


@login_required
def subject_delete(request, subject_id):
    """Fanni o'chirish"""
    
    if request.user.role != 'department_head' or not request.user.managed_department:
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    department = request.user.managed_department
    subject = get_object_or_404(Subject, id=subject_id, department=department)
    
    # Taqsimotlar borligini tekshirish
    allocations_count = subject.allocations.count()
    
    if request.method == 'POST':
        if allocations_count > 0:
            messages.error(
                request, 
                f"'{subject.name}' fanini o'chirib bo'lmaydi, chunki {allocations_count} ta taqsimot mavjud."
            )
        else:
            subject_name = subject.name
            subject.delete()
            messages.success(request, f"✓ '{subject_name}' fani o'chirildi.")
        
        return redirect('subjects_list')
    
    context = {
        'subject': subject,
        'department': department,
        'allocations_count': allocations_count,
    }
    
    return render(request, 'department_head/subject_delete_confirm.html', context)


@login_required
def allocations_list(request):
    """Fan taqsimotlari ro'yxati"""
    
    if request.user.role != 'department_head' or not request.user.managed_department:
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    department = request.user.managed_department
    
    # Filtrlar
    academic_year_id = request.GET.get('academic_year')
    semester = request.GET.get('semester')
    teacher_id = request.GET.get('teacher')
    subject_id = request.GET.get('subject')
    
    allocations = TeachingAllocation.objects.filter(
        department=department
    ).select_related('teacher', 'subject', 'group', 'academic_year')
    
    if academic_year_id:
        allocations = allocations.filter(academic_year_id=academic_year_id)
    
    if semester:
        allocations = allocations.filter(semester=semester)
    
    if teacher_id:
        allocations = allocations.filter(teacher_id=teacher_id)
    
    if subject_id:
        allocations = allocations.filter(subject_id=subject_id)
    
    allocations = allocations.order_by('-academic_year__start_date', 'semester', 'group__name')
    
    # Pagination
    paginator = Paginator(allocations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Filter options
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    teachers = User.objects.filter(active_role__role_type='teacher', department=department).order_by('first_name')
    subjects = Subject.objects.filter(department=department).order_by('name')

    import_form = AllocationImportForm()
    
    context = {
        'department': department,
        'allocations': page_obj,
        'academic_years': academic_years,
        'teachers': teachers,
        'subjects': subjects,
        'selected_year': academic_year_id,
        'selected_semester': semester,
        'selected_teacher': teacher_id,
        'selected_subject': subject_id,
        'total_count': allocations.count(),
        'import_form': import_form,
    }
    
    return render(request, 'department_head/allocations_list.html', context)


@login_required
def allocations_import(request):
    """Taqsimotlarni CSV/XLSX orqali yuklash"""
    if request.user.role != 'department_head' or not request.user.managed_department:
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')

    department = request.user.managed_department
    if request.method != 'POST':
        return redirect('allocations_list')

    form = AllocationImportForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Fayl xato: " + "; ".join(form.errors.get('file', [])))
        return redirect('allocations_list')

    alias_map = {
        'teacher': 'teacher',
        'oqituvchi': 'teacher',
        'oqituvchi_username': 'teacher',
        'username': 'teacher',
        'email': 'teacher',
        'subject_code': 'subject_code',
        'fan_kodi': 'subject_code',
        'subject_name': 'subject_name',
        'fan_nomi': 'subject_name',
        'group': 'group',
        'guruh': 'group',
        'group_name': 'group',
        'academic_year': 'academic_year',
        'oquv_yili': 'academic_year',
        'semester': 'semester',
        'semestr': 'semester',
    }

    try:
        rows = _parse_import_file(form.cleaned_data['file'])
    except Exception as e:
        messages.error(request, f"Faylni o'qishda xatolik: {str(e)}")
        return redirect('allocations_list')

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        data = _map_row_keys(row, alias_map)

        teacher_value = str(data.get('teacher') or '').strip()
        subject_code = str(data.get('subject_code') or '').strip()
        subject_name = str(data.get('subject_name') or '').strip()
        group_name = str(data.get('group') or '').strip()
        academic_year_name = str(data.get('academic_year') or '').strip()
        semester_value = str(data.get('semester') or '').strip()

        if not teacher_value or not group_name or not academic_year_name or not semester_value:
            skipped += 1
            errors.append(f"{idx}-qator: majburiy maydonlar yetarli emas.")
            continue

        try:
            semester = int(semester_value)
            if semester < 1 or semester > 8:
                raise ValueError("semestr diapazoni")
        except Exception:
            skipped += 1
            errors.append(f"{idx}-qator: semestr noto'g'ri qiymat.")
            continue

        teacher_qs = User.objects.filter(
            active_role__role_type='teacher',
            department=department
        )
        if '@' in teacher_value:
            teacher = teacher_qs.filter(email__iexact=teacher_value).first()
        else:
            teacher = teacher_qs.filter(username__iexact=teacher_value).first()
        if not teacher and ' ' in teacher_value:
            parts = teacher_value.split()
            if len(parts) >= 2:
                teacher = teacher_qs.filter(
                    first_name__iexact=parts[0],
                    last_name__iexact=parts[-1]
                ).first()
        if not teacher:
            skipped += 1
            errors.append(f"{idx}-qator: o'qituvchi topilmadi.")
            continue

        subject_qs = Subject.objects.filter(department=department)
        subject = None
        if subject_code:
            subject = subject_qs.filter(code__iexact=subject_code).first()
        if not subject and subject_name:
            subject = subject_qs.filter(name__iexact=subject_name).first()
        if not subject:
            skipped += 1
            errors.append(f"{idx}-qator: fan topilmadi.")
            continue

        group = Group.objects.filter(
            name__iexact=group_name,
            program__department=department
        ).first()
        if not group:
            skipped += 1
            errors.append(f"{idx}-qator: guruh topilmadi.")
            continue

        academic_year = AcademicYear.objects.filter(name__iexact=academic_year_name).first()
        if not academic_year:
            skipped += 1
            errors.append(f"{idx}-qator: o'quv yili topilmadi.")
            continue

        try:
            with transaction.atomic():
                allocation, created_flag = TeachingAllocation.objects.update_or_create(
                    subject=subject,
                    group=group,
                    academic_year=academic_year,
                    semester=semester,
                    defaults={
                        'teacher': teacher,
                        'department': department,
                        'created_by': request.user,
                    }
                )
            if created_flag:
                created += 1
            else:
                updated += 1
        except Exception as e:
            skipped += 1
            errors.append(f"{idx}-qator: saqlashda xatolik: {str(e)}")

    message = f"Import yakunlandi. Yangi: {created}, yangilangan: {updated}, o'tkazib yuborilgan: {skipped}."
    if errors:
        message += " Xatolar: " + " | ".join(errors[:10])
    messages.info(request, message)
    return redirect('allocations_list')


@login_required
def allocation_add(request):
    """Yangi taqsimot qo'shish"""
    
    if request.user.role != 'department_head' or not request.user.managed_department:
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    department = request.user.managed_department
    
    if request.method == 'POST':
        form = TeachingAllocationForm(request.POST, department=department)
        
        if form.is_valid():
            try:
                allocation = form.save(commit=False)
                allocation.department = department
                allocation.created_by = request.user
                
                # O'quv yilini alohida olamiz
                academic_year = form.cleaned_data['academic_year']
                allocation.academic_year = academic_year
                
                allocation.save()
                
                # Guruh ma'lumotlari
                group = allocation.group
                admission_year = None
                import re
                match = re.search(r'[-_](\d{2})$', group.name)
                if match:
                    year_suffix = match.group(1)
                    admission_year = 2000 + int(year_suffix)
                
                messages.success(
                    request,
                    f"✓ {allocation.teacher.get_full_name()} ga '{allocation.subject.name}' "
                    f"fani biriktirildi.\n"
                    f"• Guruh: {group.name} ({admission_year}-yil qabul)\n"
                    f"• O'quv yili: {academic_year.name}\n"
                    f"• Semestr: {allocation.semester}"
                )
                return redirect('allocations_list')
                
            except Exception as e:
                messages.error(request, f"Saqlashda xatolik: {str(e)}")
    else:
        form = TeachingAllocationForm(department=department)
    
    context = {
        'form': form,
        'department': department,
    }
    
    return render(request, 'department_head/allocation_form.html', context)
    


@login_required
def allocation_delete(request, allocation_id):
    """Taqsimotni o'chirish"""
    
    if request.user.role != 'department_head' or not request.user.managed_department:
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    department = request.user.managed_department
    allocation = get_object_or_404(
        TeachingAllocation, 
        id=allocation_id, 
        department=department
    )
    
    if request.method == 'POST':
        teacher_name = allocation.teacher.get_full_name()
        subject_name = allocation.subject.name
        allocation.delete()
        
        messages.success(
            request,
            f"✓ {teacher_name} dan '{subject_name}' fani olindi."
        )
        return redirect('allocations_list')
    
    context = {
        'allocation': allocation,
        'department': department,
    }
    
    return render(request, 'department_head/allocation_delete_confirm.html', context)
